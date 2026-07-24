import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


folder = Path(__file__).parent

trip_file = folder / "trip_data.json"
weather_file = folder / "Europe_Weather.xlsx"
alert_file = folder / "Europe_Weather_Alerts.txt"
dashboard_file = folder / "index.html"


# 讀取旅行資料
with open(trip_file, "r", encoding="utf-8") as file:
    trip_data = json.load(file)

alert_text = ""

if alert_file.exists():
    alert_text = alert_file.read_text(
        encoding="utf-8"
    )

if alert_text:
    alert_text = (
        alert_text
        .replace("歐洲旅行天氣提醒", "")
        .replace("====================", "")
        .strip()
    )

# 讀取天氣資料
weather_by_city = {}

if weather_file.exists():
    workbook = load_workbook(weather_file, data_only=True)
    sheet = workbook["Europe Forecast"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        city, date, min_temp, max_temp, weather, rain, advice = row

        if city not in weather_by_city:
            weather_by_city[city] = {
                "date": date,
                "min_temp": min_temp,
                "max_temp": max_temp,
                "weather": weather,
                "rain": rain,
                "advice": advice,
            }

print(city, rain, advice)
trip_name = trip_data.get("trip_name", "Europe Trip")
departure_date_text = trip_data.get("departure_date", "")

countdown_text = ""

if departure_date_text:
    departure_date = datetime.strptime(
        departure_date_text,
        "%Y-%m-%d"
    ).date()

    days_left = (
        departure_date - datetime.now().date()
    ).days

    if days_left > 0:
        countdown_text = f"距離出發還有 {days_left} 天"
    elif days_left == 0:
        countdown_text = "今天出發！"
    else:
        countdown_text = "旅程已經開始"

cities = trip_data.get("cities", [])

total_cities = len(cities)

hotel_count = sum(
    1 for city in cities
    if city.get("hotel", "").strip()
)

transport_count = sum(
    1 for city in cities
    if city.get("transport", "").strip()
)

attraction_count = sum(
    1 for city in cities
    if city.get("booked_attractions", [])
)
total_attractions = sum(
    len(city.get("booked_attractions", []))
    for city in cities
)

def make_progress_bar(done, total, color):
    if total == 0:
        percent = 0
    else:
        percent = round(done / total * 100)

    return f"""
    <div style="margin-top:4px;">
        <div style="
            width:240px;
            height:18px;
            background:#e5e7eb;
            border-radius:9px;
            overflow:hidden;
        ">
            <div style="
                width:{percent}%;
                height:100%;
                background:{color};
            "></div>
        </div>
        <div>{percent}%</div>
    </div>
    """
hotel_progress = make_progress_bar(
    hotel_count,
    total_cities,
    "#22c55e"
)

transport_progress = make_progress_bar(
    transport_count,
    total_cities,
    "#3b82f6"
)

attraction_progress = make_progress_bar(
    attraction_count,
    total_cities,
    "#f59e0b"
)

missing_hotels = [
    city.get("city", "")
    for city in cities
    if not city.get("hotel", "").strip()
]

missing_transport = [
    city.get("city", "")
    for city in cities
    if not city.get("transport", "").strip()
]

missing_attractions = [
    city.get("city", "")
    for city in cities
    if not city.get("booked_attractions", [])
]

missing_hotels_html = (
    "<br>".join(f"• {city}" for city in missing_hotels)
    if missing_hotels
    else "全部完成"
)

missing_transport_html = (
    "<br>".join(f"• {city}" for city in missing_transport)
    if missing_transport
    else "全部完成"
)

missing_attractions_html = (
    "<br>".join(f"• {city}" for city in missing_attractions)
    if missing_attractions
    else "全部完成"
)

today = datetime.now().date()

upcoming_cities = []

for city in cities:
    dates_text = city.get("dates", "").strip()

    if not dates_text or "~" not in dates_text:
        continue

    start_text = dates_text.split("~")[0].strip()

    try:
        start_date = datetime.strptime(
            start_text,
            "%Y-%m-%d"
        ).date()
    except ValueError:
        continue

    if start_date >= today:
        upcoming_cities.append((start_date, city))

upcoming_cities.sort(key=lambda item: item[0])

next_city = upcoming_cities[0][1] if upcoming_cities else None
current_city = None

for city in cities:
    dates_text = city.get("dates", "").strip()

    if "~" not in dates_text:
        continue

    start_date = datetime.strptime(
        dates_text.split("~")[0].strip(),
        "%Y-%m-%d"
    ).date()

    end_date = datetime.strptime(
        dates_text.split("~")[1].strip(),
        "%Y-%m-%d"
    ).date()

    if start_date <= today <= end_date:
        current_city = city
        break


current_city_card = f"""
<div style="
    flex:1;
    min-width:260px;
    background:#eefaf0;
    padding:14px;
    border-radius:10px;
">
<h3>📍 本站</h3>
<p>⏳ 旅程尚未開始</p>
<p>出發日期：{departure_date_text}</p>
<p>{countdown_text}</p>
</div>
"""

if current_city:
    current_city_card = f"""
<div style="
    flex:1;
    min-width:260px;
    background:#eef6ff;
    padding:14px;
    border-radius:10px;
">
<h3>📍 本站</h3>
<p><strong>{current_city.get("city", "")}, {current_city.get("country", "")}</strong></p>
<p>日期：{current_city.get("dates", "")}</p>
<p>住宿：{current_city.get("hotel", "") or "尚未填寫"}</p>
<p>交通：{current_city.get("transport", "") or "尚未填寫"}</p>
</div>
"""

next_city_weather = None

if next_city:
    next_city_name = next_city.get("city", "")
    next_city_weather = weather_by_city.get(next_city_name)
next_city_card = ""

if next_city:
    next_city_card = f"""
<div style="
    flex:1;
    min-width:260px;
    background:#eef6ff;
    padding:14px;
    border-radius:10px;
">
<h3>➡️ 下一站</h3>
<p><strong>{next_city.get("city", "")}, {next_city.get("country", "")}</strong></p>
<p>日期：{next_city.get("dates", "")}</p>
<p>
🏨 {
"已訂房" if next_city.get("hotel", "").strip()
else "尚未訂房"
}
</p>

<p>
{next_city.get("hotel", "") or "尚未填寫"}
</p>
<p>交通：{next_city.get("transport", "") or "尚未填寫"}</p>

{f"""
<p>
🌤 {next_city_weather.get('min_temp','')}
~ {next_city_weather.get('max_temp','')}°C
</p>

<p>
☁️ {next_city_weather.get('weather','')}
</p>

<p>
☂️ 降雨機率：{next_city_weather.get('rain','')}
</p>

<p>
⚠️ {next_city_weather.get('advice','')}
</p>
""" if next_city_weather else ""}

</div>
"""
html = f"""

<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">

<title>{trip_name}</title>

<style>
body {{
    font-family: Arial, sans-serif;
    background: #f4f6f8;
    margin: 0;
    padding: 20px;
}}

.container {{
    max-width: 900px;
    margin: auto;
}}

.header {{
    background: white;
    padding: 22px;
    border-radius: 14px;
    margin-bottom: 18px;
}}

.city-card {{
    background: white;
    padding: 18px;
    margin-bottom: 16px;
    border-radius: 14px;
}}

h1, h2 {{
    margin-top: 0;
}}

.section-title {{
    font-weight: bold;
    margin-top: 12px;
}}

.weather {{
    background: #eef6ff;
    padding: 12px;
    border-radius: 10px;
    margin-top: 12px;
}}

.empty {{
    color: #777;
}}
</style>
</head>

<body>
<div class="container">

<div class="header">
<h1>🌍 {trip_name}</h1>
<p>🕒 更新時間：{datetime.now().strftime("%Y-%m-%d %H:%M")}</p>
<p>{countdown_text}</p>
{f'''
<div style="
    background:#fff4d6;
    padding:14px;
    border-radius:10px;
    margin:14px 0;
">
<h3>⚠ 天氣提醒</h3>
<pre style="
    white-space:pre-wrap;
    font-family:Arial, sans-serif;
    margin:0;
">{alert_text}</pre>
</div>
''' if alert_text else ''}
<div style="
    display:flex;
    gap:12px;
    flex-wrap:wrap;
    margin:14px 0;
">
{current_city_card}
{next_city_card}
</div>

<p>
🏨 訂房進度：{hotel_count} / {total_cities}<br>
{hotel_progress}
</p>

<p>
🚆 交通進度：{transport_count} / {total_cities}<br>
{transport_progress}
</p>

<p>
🎫 景點進度：{attraction_count} / {total_cities}<br>
{attraction_progress}
</p>
<p>🎟 已填景點總數：{total_attractions}</p>

<h3>⚠ 尚未完成</h3>

<p>
🏨 尚未訂房（{len(missing_hotels)}）<br>
{missing_hotels_html}
</p>

<p>
🚆 尚未安排交通（{len(missing_transport)}）<br>
{missing_transport_html}
</p>

<p>
🎫 尚未填景點（{len(missing_attractions)}）<br>
{missing_attractions_html}
</p>

</div>
"""


for city_info in trip_data.get("cities", []):
    city = city_info.get("city", "")
    country = city_info.get("country", "")
    dates = city_info.get("dates", "")
    hotel = city_info.get("hotel", "")
    transport = city_info.get("transport", "")
    attractions = city_info.get("booked_attractions", [])
    notes = city_info.get("notes", "")

    weather = weather_by_city.get(city)

    html += f"""
    <div class="city-card">
    <h2>📍 {city}, {country}</h2>

    <div class="section-title">日期</div>
    <div>{dates if dates else '<span class="empty">尚未填寫</span>'}</div>

    <div class="section-title">住宿</div>
    <div>{hotel if hotel else '<span class="empty">尚未填寫</span>'}</div>

    <div class="section-title">交通</div>
    <div>{transport if transport else '<span class="empty">尚未填寫</span>'}</div>
    """

    if attractions:
        html += """
        <div class="section-title">已預約景點</div>
        <ul>
        """

        for attraction in attractions:
            html += f"<li>✅ {attraction}</li>"

        html += "</ul>"

    else:
        html += """
        <div class="section-title">已預約景點</div>
        <div class="empty">尚未填寫</div>
        """

    if weather:
        html += f"""
        <div class="weather">
        <strong>🌦 天氣</strong><br>
        日期：{weather["date"]}<br>
        溫度：{weather["min_temp"]}～{weather["max_temp"]}°C<br>
        天氣：{weather["weather"]}<br>
        降雨機率：{weather["rain"]}<br>
        建議：{weather["advice"]}
        </div>
        """
    else:
        html += """
        <div class="weather">
        尚未找到這個城市的天氣資料。
        </div>
        """

    if notes:
        html += f"""
        <div class="section-title">備註</div>
        <div>{notes}</div>
        """

    html += "</div>"


html += """
</div>
</body>
</html>
"""


dashboard_file.write_text(
    html,
    encoding="utf-8"
)

print("Europe Dashboard 已建立：")
print(dashboard_file)
