from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

folder = Path(__file__).parent

excel_file = folder / "Europe_Weather.xlsx"
html_file = folder / "Europe_Travel_Brief.html"

wb = load_workbook(excel_file, data_only=True)
sheet = wb["Europe Forecast"]

today = datetime.now().strftime("%Y-%m-%d")

html = f"""
<html>
<head>
<meta charset="utf-8">
<title>Europe Travel Brief</title>
</head>

<body style="font-family:Arial;max-width:900px;margin:auto">

<h1>🌍 歐洲旅行每日簡報</h1>
<h3>{today}</h3>
<hr>
"""

seen = set()

for row in sheet.iter_rows(min_row=2, values_only=True):

    city, date, min_temp, max_temp, weather, rain, advice = row

    if city in seen:
        continue

    seen.add(city)

    rain_value = int(str(rain).replace("%",""))

    alert = "✓ 適合外出"

    if rain_value >= 60:
        alert = f"☂ 降雨機率 {rain}"

    elif max_temp >= 30:
        alert = "🔥 高溫注意防曬"

    elif min_temp <= 8:
        alert = "🧥 早晚偏冷"

    html += f"""
    <div style="border:1px solid #ccc;
                padding:15px;
                margin:10px;
                border-radius:10px">

    <h2>📍 {city}</h2>

    <p>
    {min_temp} ~ {max_temp}°C
    </p>

    <p>
    {weather}
    </p>

    <b>{alert}</b>

    </div>
    """

html += """
</body>
</html>
"""

html_file.write_text(
    html,
    encoding="utf-8"
)

print("HTML 簡報已建立")