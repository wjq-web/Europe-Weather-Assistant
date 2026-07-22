from pathlib import Path

from openpyxl import load_workbook


folder = Path(__file__).parent
excel_file = folder / "Europe_Weather.xlsx"
alert_file = folder / "Europe_Weather_Alerts.txt"


if not excel_file.exists():
    print("找不到 Europe_Weather.xlsx")
    raise SystemExit(1)


workbook = load_workbook(excel_file, data_only=True)
sheet = workbook["Europe Forecast"]

alerts = []

# 從第2行開始，略過標題
for row in sheet.iter_rows(min_row=2, values_only=True):
    city, date, min_temp, max_temp, description, rain_text, advice = row

    if city is None:
        continue

    rain_chance = int(
        str(rain_text).replace("%", "").strip()
    )

    messages = []

    if rain_chance >= 60:
        messages.append(f"降雨機率 {rain_chance}%：建議帶傘")

    if max_temp is not None and max_temp >= 30:
        messages.append(f"最高溫 {max_temp}°C：注意防曬和補水")

    if min_temp is not None and min_temp <= 8:
        messages.append(f"最低溫 {min_temp}°C：注意保暖")

    if messages:
        alerts.append(
            f"{city}｜{date}\n"
            + "\n".join(f"  - {message}" for message in messages)
        )


if alerts:
    report = (
        "歐洲旅行天氣提醒\n"
        "====================\n\n"
        + "\n\n".join(alerts)
    )
else:
    report = (
        "歐洲旅行天氣提醒\n"
        "====================\n\n"
        "未發現明顯的下雨、高溫或低溫風險。"
    )


alert_file.write_text(report, encoding="utf-8")

print(report)
print("\n提醒文件已儲存：")
print(alert_file)