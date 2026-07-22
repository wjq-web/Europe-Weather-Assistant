from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

folder = Path(__file__).parent

excel_file = folder / "Europe_Weather.xlsx"
brief_file = folder / "Europe_Travel_Brief.txt"

wb = load_workbook(excel_file, data_only=True)
sheet = wb["Europe Forecast"]

today = datetime.now().strftime("%Y-%m-%d")

lines = []
lines.append("=" * 34)
lines.append("歐洲旅行每日簡報")
lines.append(today)
lines.append("=" * 34)
lines.append("")

seen_cities = set()

for row in sheet.iter_rows(min_row=2, values_only=True):

    city, date, min_temp, max_temp, weather, rain, advice = row

    if city in seen_cities:
        continue

    seen_cities.add(city)

    lines.append(city)
    lines.append(f"{min_temp} ~ {max_temp}°C")
    lines.append(weather)

    rain_value = int(str(rain).replace("%", ""))

    if rain_value >= 60:
        lines.append(f"⚠ 降雨機率 {rain}")

    elif max_temp >= 30:
        lines.append("⚠ 高溫注意防曬")

    elif min_temp <= 8:
        lines.append("⚠ 早晚偏冷")

    else:
        lines.append("✓ 適合外出")

    lines.append("")

brief_text = "\n".join(lines)

brief_file.write_text(
    brief_text,
    encoding="utf-8"
)

print(brief_text)